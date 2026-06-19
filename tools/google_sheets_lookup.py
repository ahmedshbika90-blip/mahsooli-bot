"""
Create ONE Excel workbook with a working VLOOKUP demo (standalone helper).

This version is made to be robust in Excel and Google Sheets:

  - Sheet 1: score_lookup
      grid_key, score_1_10

  - Sheet 2: farmers_demo
      farmer_id, lat, lon, grid_key, rainfall_score_vlookup, expected_score_python_check

Important:
  The grid_key in column D is written as a real value, not as a formula.
  The VLOOKUP in column E is the working formula demonstration.

Why:
  Some Excel installs do not immediately calculate formulas written by openpyxl,
  or they repair the workbook and leave formula cells blank. If D is blank,
  VLOOKUP returns #N/A. Writing D as a value makes the demo reliable.

Output:
    data/05_tables/Mahsooli_CHIRPS_..._processed_<date>.xlsx  (written locally)

Run:
  python tools/google_sheets_lookup.py

Dependency:
  python -m pip install openpyxl
"""

import argparse
import csv
import sys
from datetime import date, datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))  # repo root, for `import config` (assumes one level below root)
import config
from snap_score import OUT_OF_AOI_MESSAGE


try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    raise SystemExit(
        "Missing dependency: openpyxl\n"
        "Install it with:\n"
        "    python -m pip install openpyxl\n"
    ) from e


LOOKUP_SHEET_NAME = "score_lookup"
DEMO_SHEET_NAME = "farmers_demo"
METHODOLOGY_SHEET_NAME = "methodology"

PROCESSING_DATE = date.today().isoformat()

BASE_DOC_NAME = "Mahsooli_CHIRPS_v2_RainfallLookup_Gedaref_2014-2024"

OUT_XLSX = config.TABLE_DIR / f"{BASE_DOC_NAME}_processed_{PROCESSING_DATE}.xlsx"

def choose_output_workbook_path(preferred_path: Path) -> Path:
    """
    Pick a writable output path.

    If today's workbook is already open in Excel, Windows may lock it for writes.
    In that case, save to a timestamp-suffixed sibling file instead of failing.
    """
    if not preferred_path.exists():
        return preferred_path

    try:
        with open(preferred_path, "ab"):
            pass
        return preferred_path
    except PermissionError:
        timestamp = datetime.now().strftime("%H%M%S")
        return preferred_path.with_name(
            f"{preferred_path.stem}_{timestamp}{preferred_path.suffix}"
        )


def parse_args():
    """CLI flags for the workbook generator."""
    parser = argparse.ArgumentParser(
        description="Create the Excel rainfall-lookup workbook locally."
    )
    return parser.parse_args()


def find_latest_detailed_lookup() -> Path:
    """
    Finds the detailed lookup CSV created by 3_average.py.

    Required columns:
      - grid_key
      - lat_center
      - lon_center
      - score_1_10
    """
    candidates = sorted(config.TABLE_DIR.glob("Mahala_CHIRPS_grid_lookup_*.csv"))

    valid = []

    for path in candidates:
        try:
            with open(path, "r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                fields = set(reader.fieldnames or [])

                required = {"grid_key", "lat_center", "lon_center", "score_1_10"}
                if required.issubset(fields):
                    valid.append(path)

        except Exception:
            continue

    if not valid:
        raise FileNotFoundError(
            f"No detailed lookup CSV found in:\n"
            f"{config.TABLE_DIR}\n\n"
            f"Run this first:\n"
            f"    python 3_average.py"
        )

    return valid[-1]


def load_lookup_rows(lookup_csv: Path):
    """
    Loads useful rows from detailed lookup CSV.
    """
    rows = []

    with open(lookup_csv, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)

        required = {"grid_key", "lat_center", "lon_center", "score_1_10"}
        missing = required - set(reader.fieldnames or [])
        if missing:
            raise ValueError(f"Missing required columns in {lookup_csv}: {sorted(missing)}")

        for row in reader:
            grid_key = (row.get("grid_key") or "").strip()
            lat_center = (row.get("lat_center") or "").strip()
            lon_center = (row.get("lon_center") or "").strip()
            score = (row.get("score_1_10") or "").strip()

            if not grid_key or not lat_center or not lon_center or not score:
                continue

            rows.append({
                "grid_key": grid_key,
                "lat_center": float(lat_center),
                "lon_center": float(lon_center),
                "score_1_10": int(float(score)),
            })

    if not rows:
        raise ValueError(f"No valid rows found in lookup CSV: {lookup_csv}")

    return rows


def build_demo_rows(farmers, lookup_by_grid_key):
    """
    Build workbook demo rows from supplied farmer GPS coordinates.

    The workbook is the client-facing VLOOKUP proof, so it must exercise the same
    farmer coordinates that snap_score/pilot_check use, not synthetic grid centers.
    """
    rows = []
    for farmer in farmers:
        lat = farmer["lat"]
        lon = farmer["lon"]
        clat, clon = config.snap_to_grid_center(lat, lon)
        grid_key = config.make_grid_key(clat, clon)
        match = lookup_by_grid_key.get(grid_key)
        rows.append({
            "farmer_id": farmer["id"],
            "lat": lat,
            "lon": lon,
            "grid_key": grid_key,
            "score_1_10": match["score_1_10"] if match else "",
            "note": "" if match else OUT_OF_AOI_MESSAGE,
        })
    return rows


def style_sheet(ws):
    """
    Simple readable formatting.
    """
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D9E2F3")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=thin)

    ws.freeze_panes = "A2"

    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))

        width = min(max(max_len + 2, 10), 65)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def create_workbook(rows, demo_rows, source_lookup_csv: Path, source_farmers_csv: Path):
    """
    Creates one workbook with:
      - score_lookup
      - farmers_demo
    """
    out_xlsx = choose_output_workbook_path(OUT_XLSX)
    wb = Workbook()

    # Force Excel to recalculate formulas on open, where supported.
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass

    # -------------------------------------------------------------------------
    # Sheet 1: score_lookup
    # -------------------------------------------------------------------------
    ws_lookup = wb.active
    ws_lookup.title = LOOKUP_SHEET_NAME

    ws_lookup.append(["grid_key", "score_1_10"])

    # Deduplicate by grid_key.
    seen = set()
    for row in rows:
        if row["grid_key"] in seen:
            continue
        seen.add(row["grid_key"])

        ws_lookup.append([
            row["grid_key"],
            row["score_1_10"],
        ])

    style_sheet(ws_lookup)

    # -------------------------------------------------------------------------
    # Sheet 2: farmers_demo
    # -------------------------------------------------------------------------
    ws_demo = wb.create_sheet(DEMO_SHEET_NAME)

    ws_demo.append([
        "farmer_id",
        "lat",
        "lon",
        "grid_key",
        "rainfall_score_vlookup",
        "expected_score_python_check",
        "check_vlookup_matches_python",
    ])

    if not demo_rows:
        raise ValueError(f"No farmer rows available for workbook demo: {source_farmers_csv}")

    for excel_row_num, row in enumerate(demo_rows, start=2):
        # Use supplied farmer coordinates. grid_key is written directly as a text
        # value so the VLOOKUP demo works immediately in Excel and Google Sheets.
        lat = row["lat"]
        lon = row["lon"]
        grid_key = row["grid_key"]

        # Working VLOOKUP formula:
        # D = grid_key
        # score_lookup!A:B = lookup table with grid_key + score
        vlookup_formula = (
            f'=VLOOKUP(D{excel_row_num},\'{LOOKUP_SHEET_NAME}\'!A:B,2,FALSE)'
        )

        ws_demo.append([
            row["farmer_id"],
            lat,
            lon,
            grid_key,
            vlookup_formula,
            row["score_1_10"],
            f"=E{excel_row_num}=F{excel_row_num}",
        ])

    # Number formats
    for row_idx in range(2, ws_demo.max_row + 1):
        ws_demo[f"B{row_idx}"].number_format = "0.000000"
        ws_demo[f"C{row_idx}"].number_format = "0.000000"
        ws_demo[f"E{row_idx}"].number_format = "0"
        ws_demo[f"F{row_idx}"].number_format = "0"

    style_sheet(ws_demo)

    # Explanatory note below demo
    note_row = ws_demo.max_row + 3
    ws_demo[f"A{note_row}"] = "How the formulas work"
    ws_demo[f"A{note_row}"].font = Font(bold=True)

    ws_demo[f"A{note_row + 1}"] = (
        "Column D contains the CHIRPS 0.05 degree grid_key for each sample farmer coordinate."
    )
    ws_demo[f"A{note_row + 2}"] = (
        "Column E contains the working VLOOKUP formula that finds that grid_key in score_lookup and returns score_1_10."
    )
    ws_demo[f"A{note_row + 3}"] = (
        "Formula used in E2:"
    )
    ws_demo[f"B{note_row + 3}"] = (
        f" =VLOOKUP(D2,'{LOOKUP_SHEET_NAME}'!A:B,2,FALSE)"
    )
    ws_demo[f"A{note_row + 4}"] = (
        "Rows above use the configured farmer GPS file; for new farmer data, calculate grid_key from lat/lon using the same 0.05 degree CHIRPS grid rule, then copy the VLOOKUP formula down."
    )
    ws_demo[f"A{note_row + 5}"] = (
        f"Source detailed lookup CSV: {source_lookup_csv.name}"
    )


    # -------------------------------------------------------------------------
    # Sheet 3: methodology
    # -------------------------------------------------------------------------
    ws_method = wb.create_sheet(METHODOLOGY_SHEET_NAME)

    methodology_rows = [
        ("Item", "Value"),
        ("Data source", "CHIRPS v2.0 monthly precipitation"),
        ("Source website", "data.chc.ucsb.edu"),
        ("Area", "Gedaref bounding box, approx. lat 13.5-14.5 N, lon 35.0-36.5 E"),
        ("Grid resolution", "0.05 degrees"),
        ("Season", "June-September"),
        ("Data period", "2014-2024"),
        ("Period note", "Client-approved inclusive period; 2014-2024 contains 11 yearly seasons."),
        ("Download/processing date", PROCESSING_DATE),
        ("Scoring scale", ">=350 mm = 10; 250-349 mm = 7; 150-249 mm = 4; <150 mm = 1"),
        ("Lookup table tab", LOOKUP_SHEET_NAME),
        ("Demo tab", DEMO_SHEET_NAME),
        ("Source detailed lookup CSV", source_lookup_csv.name),
        ("Source farmer CSV", str(source_farmers_csv)),
        ("Generated workbook name", out_xlsx.name),
    ]

    for item, value in methodology_rows:
        ws_method.append([item, value])

    style_sheet(ws_method)
    ws_method.column_dimensions["A"].width = 28
    ws_method.column_dimensions["B"].width = 120
    ws_method["B11"].alignment = Alignment(wrap_text=True, vertical="top")


    # Save
    config.TABLE_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)

    return out_xlsx


def main():
    args = parse_args()

    print("=" * 90)
    print("Create single Excel workbook with lookup + VLOOKUP demo")
    print("=" * 90)

    lookup_csv = find_latest_detailed_lookup()
    print(f"Input detailed lookup: {lookup_csv}")

    rows = load_lookup_rows(lookup_csv)
    print(f"Valid lookup rows loaded: {len(rows)}")
    lookup_by_grid_key = {row["grid_key"]: row for row in rows}

    farmers_csv = config.SNAP_FARMERS_CSV
    if farmers_csv is None or not farmers_csv.exists():
        raise FileNotFoundError(
            f"Farmer CSV not found: {farmers_csv}\n"
            "Set SNAP_FARMERS_CSV to the client farmer file with id, lat, lon columns."
        )
    farmers = config.load_farmers(farmers_csv)
    if not farmers:
        raise ValueError(f"No valid farmer rows found in: {farmers_csv}")
    demo_rows = build_demo_rows(farmers, lookup_by_grid_key)
    print(f"Farmer demo rows loaded: {len(demo_rows)} from {farmers_csv}")

    out_xlsx = create_workbook(rows, demo_rows, lookup_csv, farmers_csv)

    print()
    print("=" * 90)
    print("DONE")
    print("=" * 90)
    print(f"Workbook saved: {out_xlsx}")
    print()
    print("Workbook sheets:")
    print(f"  1. {LOOKUP_SHEET_NAME}")
    print(f"  2. {DEMO_SHEET_NAME}")
    print(f"  3. {METHODOLOGY_SHEET_NAME}")
    print()
    print("Open the workbook. The VLOOKUP should return score_1_10 in farmers_demo column E.")


if __name__ == "__main__":
    main()
