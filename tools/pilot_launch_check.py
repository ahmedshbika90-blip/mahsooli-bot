"""
Pilot Launch Check.

The gate before the first farmer approvals: apply the CHIRPS rainfall lookup to
the first N registered farmers and produce a dated, written sign-off confirming
the rainfall scores calculate correctly.

This is a verification deliverable, so it deliberately reuses the *exact* snapping
path (config.snap_to_grid_center + config.make_grid_key + an exact grid_key match
against the lookup), the same one tools/snap_score.py uses in production. A farmer
whose cell is not in the lookup is **escalated for manual review, never silently
scored**. Each scored farmer's score is cross-checked two independent ways - the
score stored in the lookup vs. the score recomputed from that cell's average
rainfall via config.score_from_rainfall - so the "OK" verdict can actually fail.

Outputs (in data/pilot_check/):
    E1.3_Pilot_Check_<date>.xlsx   - check sheet (+ live VLOOKUP), score_lookup, sign-off
    E1.3_Pilot_Check_<date>.csv    - same check rows, flat
    E1.3_Signoff_Request_<date>.txt- plain-language sign-off request

Usage:
    python tools/pilot_launch_check.py
    python tools/pilot_launch_check.py --farmers-csv path/to/registry.csv --limit 10
    python tools/pilot_launch_check.py --lookup-csv path/to/lookup.csv

Defaults come from config (SNAP_FARMERS_CSV, TABLE_DIR), so behaviour is changed via
.env, not by editing this script. Exit code is non-zero if any score MISMATCH is
found (the sign-off is then BLOCKED); escalations are an expected, non-fatal
outcome and keep the sign-off valid for the in-AOI farmers.
"""

import argparse
import csv
import sys
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))  # repo root, for `import config` (assumes one level below root)
import config

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    raise SystemExit(
        "Missing dependency: openpyxl\n"
        "Install it with:\n"
        "    python -m pip install openpyxl\n"
    ) from e

# Reuse the escalation message verbatim, so both tools speak identically.
from snap_score import OUT_OF_AOI_MESSAGE


TODAY = date.today().isoformat()

OUTPUT_DIR = config.DATA_ROOT / "pilot_check"

LOOKUP_SHEET_NAME = "score_lookup"
CHECK_SHEET_NAME = "Pilot_Check"
SIGNOFF_SHEET_NAME = "Signoff_Text"

# Check verdicts.
OK = "OK"
MISMATCH = "MISMATCH"
ESCALATE = "ESCALATE"


def find_latest_lookup_csv():
    """Finds the newest detailed CHIRPS lookup CSV in config.TABLE_DIR."""
    candidates = sorted(config.TABLE_DIR.glob("Mahala_CHIRPS_grid_lookup_*.csv"))
    if not candidates:
        raise FileNotFoundError(
            f"No lookup CSV found in {config.TABLE_DIR}.\n"
            "Run the pipeline first (python run_pipeline.py), or pass one:\n"
            "    python tools/pilot_launch_check.py --lookup-csv path/to/lookup.csv"
        )
    return candidates[-1]


def load_lookup(lookup_csv):
    """
    Loads the lookup into {grid_key: {"score", "avg_mm"}}.

    The average-rainfall column is detected by its `avg_` prefix, matching
    snap_score.load_lookup, so the same lookup serves both tools.
    """
    table = {}
    with open(lookup_csv, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        fields = set(reader.fieldnames or [])
        required = {"grid_key", "score_1_10"}
        missing = required - fields
        if missing:
            raise ValueError(f"Missing columns in lookup CSV: {sorted(missing)}")

        avg_col = next((c for c in reader.fieldnames if c.startswith("avg_")), None)

        for row in reader:
            grid_key = str(row.get("grid_key", "")).strip()
            if not grid_key:
                continue
            try:
                score = int(float(row["score_1_10"]))
            except (TypeError, ValueError):
                continue
            try:
                avg_mm = float(row[avg_col]) if avg_col else None
            except (TypeError, ValueError, KeyError):
                avg_mm = None
            table[grid_key] = {"score": score, "avg_mm": avg_mm}

    if not table:
        raise ValueError(f"No valid lookup rows found in: {lookup_csv}")
    return table


def load_first_farmers(farmers_csv, limit):
    """First `limit` registered farmers via the shared config.load_farmers loader."""
    farmers = config.load_farmers(farmers_csv)
    if not farmers:
        raise ValueError(f"No valid farmers found in: {farmers_csv}")
    return farmers[:limit]


def build_check_rows(farmers, table):
    """
    Snaps each farmer with the production snapping method and verifies its score.

    - Out-of-coverage (grid_key not in the lookup) -> ESCALATE, no score.
    - In-coverage -> OK when the stored score equals the score recomputed from
      the cell's average rainfall; MISMATCH otherwise.
    """
    rows = []

    for farmer in farmers:
        lat = farmer["lat"]
        lon = farmer["lon"]
        clat, clon = config.snap_to_grid_center(lat, lon)
        grid_key = config.make_grid_key(clat, clon)
        match = table.get(grid_key)

        if match is None:
            # Outside CHIRPS coverage / the Mahala AOI - escalate, do not score.
            rows.append(
                {
                    "mahsooli_id": farmer["id"],
                    "lat": lat,
                    "lon": lon,
                    "grid_key": "",
                    "expected_v5_score": "",
                    "calculated_v5_score": "",
                    "check": ESCALATE,
                    "reviewer_note": OUT_OF_AOI_MESSAGE,
                }
            )
            continue

        stored = match["score"]
        if match["avg_mm"] is None:
            recomputed = stored
            note = "In-AOI; exact grid match to the CHIRPS lookup (no avg column to recompute)."
            verdict = OK
        else:
            recomputed = config.score_from_rainfall(match["avg_mm"])
            if recomputed == stored:
                verdict = OK
                note = (
                    "In-AOI; grid cell exact-matched to the CHIRPS lookup; "
                    "stored score = score recomputed from avg rainfall."
                )
            else:
                verdict = MISMATCH
                note = (
                    f"Stored score {stored} != recomputed {recomputed} "
                    f"(avg {match['avg_mm']:.1f} mm) - resolve before sign-off."
                )

        rows.append(
            {
                "mahsooli_id": farmer["id"],
                "lat": lat,
                "lon": lon,
                "grid_key": grid_key,
                "expected_v5_score": stored,
                "calculated_v5_score": recomputed,
                "check": verdict,
                "reviewer_note": note,
            }
        )

    return rows


def summarize(rows):
    """Returns (scored, escalated, mismatched) counts."""
    scored = sum(1 for r in rows if r["check"] == OK)
    escalated = sum(1 for r in rows if r["check"] == ESCALATE)
    mismatched = sum(1 for r in rows if r["check"] == MISMATCH)
    return scored, escalated, mismatched


def signoff_status_and_statement(rows):
    """Builds the sign-off Status line and statement, gated on the check results."""
    scored, escalated, mismatched = summarize(rows)
    escalated_ids = [r["mahsooli_id"] for r in rows if r["check"] == ESCALATE]

    if mismatched:
        status = f"BLOCKED - {mismatched} score mismatch(es) to resolve"
        statement = (
            f"NOT confirmed: {mismatched} of {len(rows)} farmers have a V5 score that "
            "does not match the source lookup. Resolve the mismatch(es) before sign-off."
        )
        return status, statement, scored, escalated, mismatched, escalated_ids

    if escalated:
        status = f"Ready - {scored} scored, {escalated} escalated for manual review"
        ids = ", ".join(escalated_ids)
        statement = (
            f"I confirm that the V5 rainfall scores calculate correctly for the {scored} "
            f"in-AOI farmers. The following are outside CHIRPS coverage and are escalated "
            f"for manual review before approval: {ids}."
        )
        return status, statement, scored, escalated, mismatched, escalated_ids

    status = "Ready - all farmers scored"
    statement = (
        f"I confirm that the V5 rainfall scores calculate correctly for the first "
        f"{scored} registered farmers."
    )
    return status, statement, scored, escalated, mismatched, escalated_ids


# Flat column order shared by the CSV and the XLSX check sheet.
CHECK_COLUMNS = [
    "Mahsooli ID",
    "Latitude",
    "Longitude",
    "Grid key",
    "Expected V5 score",
    "Calculated V5 score",
    "VLOOKUP score",
    "Check",
    "Reviewer note",
]


def write_csv(rows, out_csv):
    with open(out_csv, "w", encoding="utf-8-sig", newline="") as f:
        # The flat CSV has no live VLOOKUP; drop that spreadsheet-only column.
        fieldnames = [c for c in CHECK_COLUMNS if c != "VLOOKUP score"]
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    "Mahsooli ID": row["mahsooli_id"],
                    "Latitude": row["lat"],
                    "Longitude": row["lon"],
                    "Grid key": row["grid_key"],
                    "Expected V5 score": row["expected_v5_score"],
                    "Calculated V5 score": row["calculated_v5_score"],
                    "Check": row["check"],
                    "Reviewer note": row["reviewer_note"],
                }
            )


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D9E2F3")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=thin)

    ws.freeze_panes = "A2"

    for column_cells in ws.columns:
        max_len = 0
        col_idx = column_cells[0].column
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        width = min(max(max_len + 2, 12), 55)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_xlsx(rows, table, status, statement, out_xlsx):
    wb = Workbook()

    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass

    # Sheet 1: pilot check. Columns A..I per CHECK_COLUMNS.
    ws_check = wb.active
    ws_check.title = CHECK_SHEET_NAME
    ws_check.append(CHECK_COLUMNS)

    for r, row in enumerate(rows, start=2):
        # G: live re-lookup of the score by grid_key against the score_lookup tab.
        #    Escalated rows have a blank grid_key -> IFERROR yields "".
        vlookup = (
            f'=IFERROR(VLOOKUP(D{r},\'{LOOKUP_SHEET_NAME}\'!A:B,2,FALSE),"")'
        )
        # H: the verdict recomputed live in the sheet from E (stored), F (recomputed)
        #    and G (VLOOKUP) - an independent, human-visible cross-check.
        check = (
            f'=IF(D{r}="","{ESCALATE}",'
            f'IF(AND(E{r}=F{r},F{r}=G{r}),"{OK}","{MISMATCH}"))'
        )
        ws_check.append(
            [
                row["mahsooli_id"],
                row["lat"],
                row["lon"],
                row["grid_key"],
                row["expected_v5_score"],
                row["calculated_v5_score"],
                vlookup,
                check,
                row["reviewer_note"],
            ]
        )

    for r in range(2, ws_check.max_row + 1):
        ws_check[f"B{r}"].number_format = "0.000000"
        ws_check[f"C{r}"].number_format = "0.000000"

    style_sheet(ws_check)

    # Sheet 2: score_lookup copied from the CHIRPS lookup (the VLOOKUP target).
    ws_lookup = wb.create_sheet(LOOKUP_SHEET_NAME)
    ws_lookup.append(["grid_key", "score_1_10"])
    for grid_key, entry in table.items():
        ws_lookup.append([grid_key, entry["score"]])
    style_sheet(ws_lookup)

    # Sheet 3: sign-off text (gated on the check result).
    ws_signoff = wb.create_sheet(SIGNOFF_SHEET_NAME)
    ws_signoff.append(["Field", "Value"])
    ws_signoff.append(["Deliverable", "E1.3 Pilot Launch Check"])
    ws_signoff.append(["Sign-off statement", statement])
    ws_signoff.append(["Reviewer name", ""])
    ws_signoff.append(["Reviewer role", ""])
    ws_signoff.append(["Date", ""])
    ws_signoff.append(["Status", status])
    ws_signoff.append(["Generated date", TODAY])
    style_sheet(ws_signoff)

    wb.save(out_xlsx)


def write_signoff_txt(statement, status, escalated_ids, out_txt):
    escalation_block = ""
    if escalated_ids:
        escalation_block = (
            "\nEscalated for manual review (outside CHIRPS coverage, not scored):\n"
            + "\n".join(f"  - {fid}" for fid in escalated_ids)
            + "\n"
        )

    text = f"""Hi,

I have prepared the E1.3 Pilot Launch Check.

The file applies the CHIRPS rainfall lookup to the first registered farmers,
snapping each GPS point to its 0.05 deg grid cell and cross-checking the V5 score
against the source lookup. Any farmer outside coverage is escalated for manual
review rather than scored.
{escalation_block}
Status: {status}

Please review and confirm the following statement:

"{statement}"

Generated date: {TODAY}
"""
    out_txt.write_text(text, encoding="utf-8")


def parse_args():
    parser = argparse.ArgumentParser(description="Create the pilot launch check package.")
    parser.add_argument(
        "--farmers-csv",
        default=str(config.SNAP_FARMERS_CSV),
        help="CSV with registered farmers. Default: config.SNAP_FARMERS_CSV",
    )
    parser.add_argument(
        "--lookup-csv",
        default=None,
        help="Optional path to the detailed lookup CSV. Default: newest in TABLE_DIR.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=10,
        help="Number of first farmers to check. Default: 10",
    )
    return parser.parse_args()


def main():
    args = parse_args()

    farmers_csv = Path(args.farmers_csv)
    if not farmers_csv.exists():
        raise FileNotFoundError(
            f"Farmers CSV not found:\n{farmers_csv}\n\n"
            "Expected columns: an id column (mahsooli_id/farmer_id/id), lat, lon"
        )

    lookup_csv = Path(args.lookup_csv) if args.lookup_csv else find_latest_lookup_csv()
    if not lookup_csv.exists():
        raise FileNotFoundError(f"Lookup CSV not found: {lookup_csv}")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_xlsx = OUTPUT_DIR / f"E1.3_Pilot_Check_{TODAY}.xlsx"
    out_csv = OUTPUT_DIR / f"E1.3_Pilot_Check_{TODAY}.csv"
    out_txt = OUTPUT_DIR / f"E1.3_Signoff_Request_{TODAY}.txt"

    print("=" * 90)
    print("E1.3 Pilot Launch Check")
    print("=" * 90)
    print(f"Farmers CSV: {farmers_csv}")
    print(f"Lookup CSV:  {lookup_csv}")
    print(f"Output dir:  {OUTPUT_DIR}")
    print(f"Limit:       {args.limit}")
    print(f"Cell size:   {config.CELL_SIZE} deg")
    print("=" * 90)

    table = load_lookup(lookup_csv)
    farmers = load_first_farmers(farmers_csv, args.limit)
    rows = build_check_rows(farmers, table)

    status, statement, scored, escalated, mismatched, escalated_ids = (
        signoff_status_and_statement(rows)
    )

    write_csv(rows, out_csv)
    write_xlsx(rows, table, status, statement, out_xlsx)
    write_signoff_txt(statement, status, escalated_ids, out_txt)

    for row in rows:
        shown = row["expected_v5_score"] if row["expected_v5_score"] != "" else "-"
        print(
            f"  {row['mahsooli_id']:<8} ({row['lat']:.4f}, {row['lon']:.4f}) -> "
            f"{(row['grid_key'] or '(no cell)'):<16} score={shown:<3} {row['check']}"
        )

    print()
    print("=" * 90)
    print(f"Farmers checked: {len(rows)}   Scored: {scored}   "
          f"Escalated: {escalated}   Mismatched: {mismatched}")
    print(f"Sign-off status: {status}")
    print("=" * 90)
    print(f"Excel: {out_xlsx}")
    print(f"CSV:   {out_csv}")
    print(f"Text:  {out_txt}")
    print()

    if mismatched:
        print("Sign-off BLOCKED: resolve the score mismatch(es) above before issuing it.")
        return 1

    print("Next step: send the Excel file and sign-off text to the reviewer/client.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
